# Katip Celebi
# Copyright (C) 2026 farukylmz0550
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

"""Spreadsheets: your library on the way out, and a list of ISBNs on
the way in.

The export is for reading elsewhere -- printing it, mailing it, keeping a copy.
It is not a second save format: the app's own files are the JSON next door, and
nothing here is ever read back into them.
"""

from pathlib import Path
from typing import Optional
import logging

try:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - openpyxl ships with the app
    openpyxl = None

from books import tags
from books.model import Book, is_valid_isbn, normalize_isbn
from books.reading import parse_stamp, reading_days, status_of
from shared.texts import field_label, text

logger = logging.getLogger("katipcelebi")

TEMPLATE_DEFAULT_NAME = "isbn_list.xlsx"
EXPORT_DEFAULT_NAME = "my_library.xlsx"

# Fields holding a timestamp. Excel gets a real datetime for these, so the
# column sorts and filters as a date instead of being text that looks like one.
DATE_FIELDS = ("started_date", "finished_date")

# The record, in the order a person would read it. `key` is not among them: it
# is the app's bookkeeping, and `isbn` below says the true thing instead.
EXPORTED_FIELDS = (
    "title",
    "subtitle",
    "authors",
    "publishers",
    "publish_date",
    "publish_places",
    "edition_name",
    "series",
    "number_of_pages",
    "languages",
    "isbn_10",
    "isbn_13",
    "subjects",
    "rating",
    "status",
    "tags",
    "signed",
    "copies",
    "notes",
    "started_date",
    "finished_date",
)

# Two columns are not fields of a book: they are worked out. Named here with
# the rest so that the formatting below can ask what a column is, rather than
# count how far along it has got.
DAYS_COLUMN = "days"
LENT_COLUMN = "lent_to"
COLUMN_NAMES = ("isbn",) + EXPORTED_FIELDS + (DAYS_COLUMN, LENT_COLUMN)


def _stamp(value: str):
    """A stored timestamp as something Excel understands as a date."""
    return parse_stamp(value) or value or ""


def export_library(books: list[Book], ledger, path: Path) -> bool:
    """Write the whole library to a spreadsheet. False when it could not be."""
    if openpyxl is None:
        logger.error("openpyxl is not installed; cannot export")
        return False

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = text("nav_library")[:31]

    headings = [field_label("isbn")] + [
        field_label(name) for name in EXPORTED_FIELDS
    ]
    headings += [text("export_days"), text("export_lent_to")]
    for column, heading in enumerate(headings, start=1):
        cell = sheet.cell(row=1, column=column, value=heading)
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"

    for row, entry in enumerate(books, start=2):
        values = [entry.display_isbn]
        for name in EXPORTED_FIELDS:
            values.append(_column_value(entry, name))
        # A number, not "2 day(s) 6 hour(s)": this is a spreadsheet, and the
        # whole reason to open it there is to sort and average a column. The
        # phrase reads better and does neither.
        days = reading_days(entry)
        values.append(round(days, 2) if days else "")
        # Every name, not the first: a book you own two of can be out
        # with two people, and naming one of them is naming the wrong one.
        out = ledger.open_loans_for(entry.key) if ledger is not None else []
        values.append(", ".join(loan.person_name for loan in out))

        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            name = COLUMN_NAMES[column - 1]
            if name in ("isbn", "isbn_10", "isbn_13"):
                # Text, or Excel renders a 13-digit ISBN as 9.78031E+12.
                cell.number_format = "@"
            elif name == DAYS_COLUMN and value != "":
                cell.number_format = "0.0"
            elif name in DATE_FIELDS and value:
                cell.number_format = "yyyy-mm-dd hh:mm"

    _widen(sheet, headings)
    try:
        book.save(path)
        logger.info("Exported %d book(s) to %s", len(books), path)
        return True
    except (OSError, ValueError):
        logger.exception("Could not write %s", path)
        return False


def _column_value(entry: Book, name: str):
    if name in DATE_FIELDS:
        return _stamp(getattr(entry, name))
    if name == "status":
        return text("status_" + status_of(entry))
    if name == "tags":
        # The words as the app has been showing them, not as it files them.
        return tags.show(entry.tags)
    if name == "signed":
        return (
            text("export_yes") if entry.signed.strip() else text("export_no")
        )
    return getattr(entry, name)


def _widen(sheet, headings: list[str]) -> None:
    """Columns wide enough to read, but not so wide the sheet is unusable."""
    for column, heading in enumerate(headings, start=1):
        longest = len(str(heading))
        for cell in sheet[get_column_letter(column)]:
            longest = max(longest, len(str(cell.value or "")))
        sheet.column_dimensions[get_column_letter(column)].width = min(
            48, max(10, longest + 2)
        )


# ------------------------------------------------------- the way back in ---
def write_template(path: Path) -> bool:
    """An empty sheet with one column, for the user to paste ISBNs into."""
    if openpyxl is None:
        return False
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = field_label("isbn")
    cell = sheet.cell(row=1, column=1, value=field_label("isbn"))
    cell.font = Font(bold=True)
    sheet.column_dimensions["A"].width = 22
    # Text, or Excel turns a pasted 13-digit ISBN into 9.78045E+12 and the app
    # would be handed a number that is no longer an ISBN.
    for row in range(2, 500):
        sheet.cell(row=row, column=1).number_format = "@"
    try:
        book.save(path)
        return True
    except (OSError, ValueError):
        logger.exception("Could not write %s", path)
        return False


def read_template(path: Path) -> Optional[list[str]]:
    """The ISBNs in a template. None when the file cannot be read at all.

    Anything that is not an ISBN is left out rather than passed on: the point
    of reading the file is to find the ISBNs in it, and a heading or a stray
    note is not a book that could not be found.
    """
    if openpyxl is None:
        return None
    try:
        book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        logger.exception("Could not read %s", path)
        return None

    found, seen = [], set()
    try:
        for sheet in book.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    if value is None:
                        continue
                    # A cell Excel decided was a number comes back as a float.
                    raw = (
                        str(int(value))
                        if isinstance(value, float) and value.is_integer()
                        else str(value)
                    )
                    isbn = normalize_isbn(raw)
                    if is_valid_isbn(isbn) and isbn not in seen:
                        seen.add(isbn)
                        found.append(isbn)
    finally:
        book.close()
    return found
